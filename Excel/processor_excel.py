import os
from openpyxl import load_workbook, Workbook
from pathlib import Path
from File_System.sort_files import iteration

"""
A script that reads all Excel files and processes them
"""

files_excel = Path('../File_System/raw_data/excel_data')


def read_all_xlsx_data(file_path):
    workbook = load_workbook(file_path)
    all_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        all_data[sheet_name] = data

    return all_data


def delete_empty():
    path = '../File_System/raw_data/excel_data.xlsx'
    wb = load_workbook(path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # We create a list to store the numbers of lines that need to be deleted
        rows_to_delete = []

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # We check if there is at least one cell with the value None
            if any(cell_value is None for cell_value in row):
                rows_to_delete.append(row_index)

        # Remove lines that are empty
        for row_index in reversed(rows_to_delete):
            sheet.delete_rows(row_index)

    # Save the changes to the file
    save_as_a_new_file(wb, 'processed_data')


def save_as_a_new_file(wb, name_file):
    # Delete the sheet "Sheet"
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    new_wb = Workbook()

    for sheet_name in wb.sheetnames:
        old_sheet = wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)

        for rows in old_sheet.iter_rows(values_only=True):
            new_sheet.append(rows)

    if not Path('../processed_data/').exists():
        os.mkdir('../processed_data/')

    new_path = f'../processed_data/{name_file}.xlsx'
    new_wb.save(new_path)
    new_wb.close()


fil = iteration(files_excel.glob("*.xlsx"))

delete_empty()
